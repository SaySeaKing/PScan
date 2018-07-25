# -*- coding:utf-8 -*-

import datetime
import openpyxl
import os

def sheet_Exist(wb, sheet_name):
    try:
        ws = wb.get_sheet_by_name(sheet_name)
        return True
    except:
        return False

def open_Excel(filename, sheetname):
    try:
        wb = openpyxl.load_workbook(filename)
    except:
        wb = openpyxl.Workbook()
    if sheet_Exist(wb, sheetname):
        ws = wb.get_sheet_by_name(sheetname)
    else:
        wb.create_sheet(sheetname)
        try:
            wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        except:
            pass
        ws = wb.get_sheet_by_name(sheetname)
    return wb, ws
        
def result_Record(ip_subnet, ip):
    time_ = datetime.datetime.now().strftime('%H:%M')
    if os.path.exists('log_ip/%s.xlsx'  % ip_subnet.split('/')[0]):
        if '00:00' <= time_ <= '00:05':
            os.remove('log_ip/%s.xlsx'  % ip_subnet.split('/')[0])
    filename = 'log_ip/%s.xlsx'  % ip_subnet.split('/')[0]
    wb, ws = open_Excel(filename, u"log")
    ws.append([ip])
    wb.save(filename)
    wb.close()

def ip_Decrease(ip_subnet, all_ip):
    IP = []
    filename = 'log_ip/%s.xlsx' % ip_subnet.split('/')[0]
    wb, ws = open_Excel(filename, u"log")
    for ip in all_ip:
        if not ip_exists(ws, ip):
            IP.append(ip)
    return IP

def ip_exists(ws, ip):
    for row in ws.rows:
        if ip == row[0].value:
            return True
    return False
