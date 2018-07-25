# -*- coding:utf-8 -*-

import datetime
import openpyxl
import os

def sheet_Exist(wb, sheet_name):
    try:
        ws = wb.get_sheet_by_name(sheet_name)
        return True
    except:
        return False

def open_Excel(filename, sheetname):
    try:
        wb = openpyxl.load_workbook(filename)
    except:
        wb = openpyxl.Workbook()
    if sheet_Exist(wb, sheetname):
        ws = wb.get_sheet_by_name(sheetname)
    else:
        wb.create_sheet(sheetname)
        ws = wb.get_sheet_by_name(sheetname)
    try:
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    except:
        pass
    return wb, ws
        
def result_Record(ip_subnet, result):
    date_ = datetime.datetime.now().strftime('%Y%m%d')
    filename = "logFile/%s_log.xlsx" % (date_)
    wb, ws = open_Excel(filename, u"log")
    port_ = ''
    for port in result[1]:
        port_ += str(port) + ','
    ws.append([result[0], port_])
    wb.save(filename)
    wb.close()
